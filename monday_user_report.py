#!/usr/bin/env python3
"""
monday_user_report.py
Python port of monday_user_report.ps1

Generates the Coral User Report — a CSV/HTML/XLSX mapping of Monday.com
workspaces to their members with role, status, team, and activity data.

Usage:
    pip install requests python-dotenv openpyxl truststore
    python monday_user_report.py

Requires a .env file in the same directory:
    MONDAY_API_TOKEN=<your_token>
"""

import csv
import os
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import requests
from dotenv import load_dotenv

# Use the Windows certificate store so corporate SSL inspection certs are trusted
try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass  # fall back to default certifi bundle

# ── Config ─────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
load_dotenv(SCRIPT_DIR / ".env")

API_URL  = "https://api.monday.com/v2"
API_VER  = "2024-07"
TOKEN    = os.getenv("MONDAY_API_TOKEN")

if not TOKEN:
    print("Error: Set MONDAY_API_TOKEN in .env first", file=sys.stderr)
    sys.exit(1)

DATE        = datetime.now().strftime("%Y-%m-%d")
REPORTS_DIR = SCRIPT_DIR / "Reports"
REPORTS_DIR.mkdir(exist_ok=True)
OUTPUT_CSV  = REPORTS_DIR / f"coral_user_report_{DATE}.csv"
OUTPUT_HTML = REPORTS_DIR / f"coral_user_report_{DATE}.html"
OUTPUT_XLSX = REPORTS_DIR / f"coral_user_report_{DATE}.xlsx"

CSV_COLUMNS = [
    "Workspace", "Name", "Email", "User Role", "Status",
    "Teams", "Joined", "Last Active", "Invited By", "2FA", "Workspace URL",
]

_HEADERS = {
    "Content-Type":  "application/json",
    "Authorization": TOKEN,
    "API-Version":   API_VER,
}


# ── Helpers ────────────────────────────────────────────────────
def run_gql(query: str) -> dict:
    resp = requests.post(API_URL, headers=_HEADERS, json={"query": query})
    resp.raise_for_status()
    return resp.json()


def get_role(member: dict) -> str:
    if member.get("is_admin"):     return "Admin"
    if member.get("is_guest"):     return "Guest"
    if member.get("is_view_only"): return "Viewer"
    return "Member"


def get_status(member: dict) -> str:
    return "Active" if member.get("enabled") else "Inactive"


def get_teams(member: dict) -> str:
    teams = member.get("teams") or []
    names = [t["name"] for t in teams if t.get("name")]
    return "; ".join(names) if names else "No Teams"


def parse_date(s: str) -> datetime | None:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s[:19])
    except ValueError:
        return None


def fmt_date(s: str) -> str:
    return s[:10] if s and len(s) >= 10 else (s or "")


def esc(s: str) -> str:
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ── Step 1: Fetch all workspaces ───────────────────────────────
print("Fetching workspaces...")

ws_result = run_gql("""
query {
  workspaces(limit: 100) {
    id
    name
    kind
  }
}
""")

if "errors" in ws_result:
    print(f"API Error fetching workspaces: {ws_result['errors']}", file=sys.stderr)
    sys.exit(1)

workspaces = ws_result["data"]["workspaces"]
print(f"Found {len(workspaces)} workspace(s).")

# ── Step 2 & 3: Iterate workspaces, collect rows ───────────────
rows: list[dict] = []

for ws in workspaces:
    ws_id   = ws["id"]
    ws_name = ws["name"]
    ws_kind = ws["kind"]   # "open" or "closed"
    ws_url  = f"https://coral.monday.com/workspaces/{ws_id}"

    print(f"  -> Processing workspace: {ws_name} (id: {ws_id}, kind: {ws_kind})")

    all_members: list[dict] = []
    page      = 1
    page_size = 100

    while True:
        if ws_kind == "open":
            # Open workspaces include all account users implicitly
            query = f"""
query {{
  users(limit: {page_size}, page: {page}, kind: all) {{
    id name email is_admin is_guest enabled created_at last_activity is_view_only
    teams {{ name }}
  }}
}}"""
            result  = run_gql(query)
            members = ((result.get("data") or {}).get("users")) or []
        else:
            # Closed workspaces use explicit subscriber list
            query = f"""
query {{
  workspaces(ids: [{ws_id}]) {{
    members: users_subscribers(limit: {page_size}, page: {page}) {{
      id name email is_admin is_guest enabled created_at last_activity is_view_only
      teams {{ name }}
    }}
  }}
}}"""
            result  = run_gql(query)
            ws_data = (((result.get("data") or {}).get("workspaces")) or [{}])[0]
            members = ws_data.get("members") or []

        if "errors" in result:
            print(f"    Warning: API error on page {page}: {result['errors']}")
            break

        count = len(members)
        print(f"    Page {page}: {count} member(s)")
        all_members.extend(members)
        page += 1
        time.sleep(0.15)

        if count < page_size:
            break

    if not all_members:
        print("    (no members found)")
        time.sleep(0.3)
        continue

    print(f"    Total members in '{ws_name}': {len(all_members)}")

    for m in all_members:
        rows.append({
            "Workspace":     ws_name,
            "Name":          m.get("name") or "",
            "Email":         m.get("email") or "",
            "User Role":     get_role(m),
            "Status":        get_status(m),
            "Teams":         get_teams(m),
            "Joined":        m.get("created_at") or "",
            "Last Active":   m.get("last_activity") or "Never logged in",
            "Invited By":    "N/A",
            "2FA":           "Disabled",
            "Workspace URL": ws_url,
        })

    time.sleep(0.3)

# ── Step 4: Strip deleted-member rows, write CSV ──────────────
before_count = len(rows)
rows = [r for r in rows if not any(v == "Deleted member" for v in r.values())]
removed_count = before_count - len(rows)

with open(OUTPUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    writer.writerows(rows)

print(f"\nDone! Report saved to: {OUTPUT_CSV}")
print(f"   Rows written: {len(rows)}")
print(f"   Deleted-member rows removed: {removed_count}")

# ── Step 5: HTML executive report ─────────────────────────────
print("Generating HTML executive report...")

cutoff = datetime.now() - timedelta(days=30)

inactive_users = sorted(
    [r for r in rows if r["Last Active"] == "Never logged in"
     or (parse_date(r["Last Active"]) and parse_date(r["Last Active"]) < cutoff)],
    key=lambda r: (r["Workspace"], r["Last Active"]),
)

guest_users = sorted(
    [r for r in rows if r["User Role"] == "Guest"],
    key=lambda r: r["Workspace"],
)

new_hires = sorted(
    [r for r in rows if parse_date(r["Joined"]) and parse_date(r["Joined"]) > cutoff],
    key=lambda r: r["Joined"],
    reverse=True,
)

ws_groups: dict[str, list] = defaultdict(list)
for r in rows:
    ws_groups[r["Workspace"]].append(r)

# Workspace breakdown rows
ws_rows_html = []
for ws_name in sorted(ws_groups):
    grp      = ws_groups[ws_name]
    ws_url   = grp[0]["Workspace URL"]
    total    = len(grp)
    admins   = sum(1 for r in grp if r["User Role"] == "Admin")
    members  = sum(1 for r in grp if r["User Role"] == "Member")
    viewers  = sum(1 for r in grp if r["User Role"] == "Viewer")
    guests   = sum(1 for r in grp if r["User Role"] == "Guest")
    active   = sum(1 for r in grp if r["Status"] == "Active")
    inactive = sum(1 for r in grp if r["Status"] == "Inactive")
    in_cls   = " class='warn'" if inactive > 0 else ""
    ws_rows_html.append(f"""    <tr>
      <td><a href="{ws_url}" target="_blank">{esc(ws_name)}</a></td>
      <td class="num">{total}</td>
      <td class="num">{admins}</td>
      <td class="num">{members}</td>
      <td class="num">{viewers}</td>
      <td class="num">{guests}</td>
      <td class="num active">{active}</td>
      <td class="num"{in_cls}>{inactive}</td>
    </tr>""")

# Inactive user rows
inactive_rows_html = []
for u in inactive_users:
    la      = u["Last Active"]
    la_cell = "<span class='never'>Never logged in</span>" if la == "Never logged in" else fmt_date(la)
    inactive_rows_html.append(f"""    <tr>
      <td>{esc(u["Name"])}</td>
      <td>{esc(u["Email"])}</td>
      <td>{esc(u["Workspace"])}</td>
      <td>{esc(u["User Role"])}</td>
      <td>{la_cell}</td>
    </tr>""")

# Guest user rows
guest_rows_html = []
for u in guest_users:
    st_cls = "active" if u["Status"] == "Active" else "warn"
    guest_rows_html.append(f"""    <tr>
      <td>{esc(u["Name"])}</td>
      <td>{esc(u["Email"])}</td>
      <td>{esc(u["Workspace"])}</td>
      <td class="{st_cls}">{esc(u["Status"])}</td>
      <td>{fmt_date(u["Joined"])}</td>
    </tr>""")

# New hire rows + workspace filter options
new_hire_workspaces = sorted({r["Workspace"] for r in new_hires})
ws_options_html = "\n".join(
    f'    <option value="{esc(w)}">{esc(w)}</option>' for w in new_hire_workspaces
)

new_hire_rows_html = []
for u in new_hires:
    new_hire_rows_html.append(f"""    <tr class="new-hire" data-workspace="{esc(u['Workspace'])}" data-status="{esc(u['Status'])}">
      <td>{esc(u["Name"])}</td>
      <td>{esc(u["Email"])}</td>
      <td>{esc(u["Workspace"])}</td>
      <td>{esc(u["User Role"])}</td>
      <td>{esc(u["Status"])}</td>
      <td>{fmt_date(u["Joined"])}</td>
      <td>{esc(u["Invited By"])}</td>
    </tr>""")


def html_section(title: str, head: str, body_rows: list, empty_msg: str) -> str:
    if body_rows:
        inner = f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"
    else:
        inner = f"<p class='empty'>{empty_msg}</p>"
    return f"<section><h2>{title}</h2>{inner}</section>"


ws_section = html_section(
    "Workspace Breakdown",
    "<th>Workspace</th><th>Total</th><th>Admins</th><th>Members</th>"
    "<th>Viewers</th><th>Guests</th><th>Active</th><th>Inactive</th>",
    ws_rows_html, "No workspace data.",
)
inactive_section = html_section(
    "Inactive Users (30+ days without login)",
    "<th>Name</th><th>Email</th><th>Workspace</th><th>Role</th><th>Last Active</th>",
    inactive_rows_html, "No inactive users found.",
)
guest_section = html_section(
    "Guest / External Users",
    "<th>Name</th><th>Email</th><th>Workspace</th><th>Status</th><th>Joined</th>",
    guest_rows_html, "No guest users found.",
)

if new_hire_rows_html:
    new_hire_inner = f"""<div class="filter-bar">
  <label for="ws-filter">Workspace</label>
  <select id="ws-filter" onchange="filterNewHires()">
    <option value="">All Workspaces</option>
    {ws_options_html}
  </select>
  <label for="st-filter">Status</label>
  <select id="st-filter" onchange="filterNewHires()">
    <option value="">All</option>
    <option value="Active">Active</option>
    <option value="Inactive">Inactive</option>
  </select>
  <span class="filter-count" id="hire-count">{len(new_hires)} hire(s)</span>
</div>
<table id="new-hire-table">
  <thead><tr><th>Name</th><th>Email</th><th>Workspace</th><th>Role</th><th>Status</th><th>Joined</th><th>Invited By</th></tr></thead>
  <tbody>{"".join(new_hire_rows_html)}</tbody>
</table>"""
else:
    new_hire_inner = "<p class='empty'>No new hires in the last 30 days.</p>"

new_hire_section = f"<section><h2>New Hires &mdash; Last 30 Days</h2>{new_hire_inner}</section>"

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Coral Monday.com User Report - {DATE}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f5f7; color: #172b4d; }}
  header {{ background: #0052cc; color: #fff; padding: 24px 40px; }}
  header h1 {{ font-size: 22px; font-weight: 600; }}
  header p  {{ font-size: 13px; opacity: 0.75; margin-top: 5px; }}
  main {{ padding: 32px 40px; max-width: 1200px; }}
  section {{ background: #fff; border-radius: 6px; box-shadow: 0 1px 3px rgba(0,0,0,.12); margin-bottom: 28px; overflow: hidden; }}
  section h2 {{ font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: .06em;
               color: #5e6c84; padding: 14px 20px; border-bottom: 1px solid #ebecf0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #f4f5f7; color: #5e6c84; font-weight: 700; text-transform: uppercase;
       font-size: 11px; letter-spacing: .05em; padding: 10px 16px; text-align: left; }}
  td {{ padding: 9px 16px; border-top: 1px solid #ebecf0; vertical-align: middle; }}
  tr:hover td {{ background: #fafbfc; }}
  td.num {{ text-align: center; }}
  .active {{ color: #006644; font-weight: 600; }}
  .warn   {{ color: #bf2600; font-weight: 600; }}
  .never  {{ color: #bf2600; }}
  a {{ color: #0052cc; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .empty {{ color: #7a869a; font-style: italic; padding: 16px 20px; font-size: 13px; }}
  .new-hire td {{ background: #e8f0fe; }}
  tr.new-hire:hover td {{ background: #d2e3fc; }}
  .filter-bar {{ padding: 12px 20px; border-bottom: 1px solid #ebecf0; display: flex; align-items: center; gap: 12px; }}
  .filter-bar label {{ font-size: 12px; font-weight: 700; color: #5e6c84; text-transform: uppercase; letter-spacing: .05em; }}
  .filter-bar select {{ font-size: 13px; padding: 5px 10px; border: 1px solid #dfe1e6; border-radius: 4px; color: #172b4d; cursor: pointer; }}
  .filter-count {{ font-size: 12px; color: #5e6c84; }}
</style>
</head>
<body>
<header>
  <h1>Coral &mdash; Monday.com User Report</h1>
  <p>Generated {DATE} &nbsp;&bull;&nbsp; {len(rows)} users across {len(ws_groups)} workspaces &nbsp;&bull;&nbsp; Source: {OUTPUT_CSV}</p>
</header>
<main>
{new_hire_section}
{ws_section}
{inactive_section}
{guest_section}
</main>
<script>
function filterNewHires() {{
  var ws = document.getElementById('ws-filter').value;
  var st = document.getElementById('st-filter').value;
  var tableRows = document.querySelectorAll('#new-hire-table tbody tr');
  var visible = 0;
  tableRows.forEach(function(row) {{
    var show = (!ws || row.dataset.workspace === ws) &&
               (!st || row.dataset.status === st);
    row.style.display = show ? '' : 'none';
    if (show) visible++;
  }});
  document.getElementById('hire-count').textContent = visible + ' hire(s)';
}}
</script>
</body>
</html>"""

OUTPUT_HTML.write_text(html, encoding="utf-8")
print(f"   HTML report saved to: {OUTPUT_HTML}")

# ── Step 6: Excel report ───────────────────────────────────────
print("Generating Excel report...")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb       = openpyxl.Workbook()
    ws_sheet = wb.active
    ws_sheet.title = "Users"

    # Header row
    ws_sheet.append(CSV_COLUMNS)
    for cell in ws_sheet[1]:
        cell.font = Font(bold=True)

    ws_sheet.auto_filter.ref = f"A1:{get_column_letter(len(CSV_COLUMNS))}1"
    ws_sheet.freeze_panes    = "A2"

    blue_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for row in rows:
        ws_sheet.append([row[c] for c in CSV_COLUMNS])
        row_idx    = ws_sheet.max_row
        joined_val = row.get("Joined") or ""
        d          = parse_date(joined_val)
        if d and d > cutoff:
            for col_idx in range(1, len(CSV_COLUMNS) + 1):
                ws_sheet.cell(row=row_idx, column=col_idx).fill = blue_fill

    # Approximate auto-size
    for col_cells in ws_sheet.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws_sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 50)

    wb.save(OUTPUT_XLSX)
    print(f"   Excel report saved to: {OUTPUT_XLSX}")

except ImportError:
    print("   Skipped — install openpyxl: pip install openpyxl")
except Exception as e:
    print(f"   Excel generation failed: {e}")
